"""
conftest_reporter.py
====================
Xuất Excel kết quả test theo đúng format file mẫu:
TEST CASE ID | TEST SCENARIO | TEST CASE | PRE-CONDITION | TEST STEPS
| TEST DATA | EXPECTED RESULT | POST CONDITION | ACTUAL RESULT | STATUS
"""

import re
import os
import datetime
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────
# Màu sắc
# ──────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1F4E79"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_PASS_BG     = "E2EFDA"
COLOR_FAIL_BG     = "FCE4D6"
COLOR_META_BG     = "D6E4F0"
COLOR_ROW_ALT     = "F2F2F2"


# ──────────────────────────────────────────────────────────────
# Map tên file → prefix TC ID
# ──────────────────────────────────────────────────────────────
FILE_PREFIX_MAP = {
    "test_01_login":    "Login",
    "test_02_register": "Register",
    "test_03_sanpham":  "SanPham",
    "test_04_giohang":  "GioHang",
    "test_05_admin":    "Admin",
    "test_06_ui_flow":  "UIFlow",
}


def _get_prefix(nodeid):
    file_part = nodeid.split("::")[0]
    filename  = os.path.basename(file_part)
    stem      = filename.replace(".py", "")
    return FILE_PREFIX_MAP.get(stem, "")


def _tc_id(prefix, index):
    if prefix:
        return f"TC_{prefix.upper()}_{index:03d}"
    return f"TC_{index:03d}"


# ──────────────────────────────────────────────────────────────
# Sinh nội dung tự động từ tên test
# ──────────────────────────────────────────────────────────────
def _make_steps(testcase, params):
    """Sinh test steps cơ bản từ tên hàm test."""
    name = testcase.replace("test_", "").replace("_", " ")
    steps = f"1. Mở trình duyệt, truy cập ứng dụng\n2. Thực hiện: {name}"
    if params:
        steps += f"\n3. Dữ liệu đầu vào: {params}"
    steps += "\n4. Quan sát kết quả"
    return steps

def _make_precondition(scenario):
    """Sinh pre-condition từ tên class."""
    s = scenario.lower()
    if "login" in s or "dang_nhap" in s:
        return "Website đang hoạt động, trình duyệt đã mở"
    if "logout" in s or "dang_xuat" in s:
        return "Đã đăng nhập thành công"
    if "admin" in s or "nhanvien" in s:
        return "Trang admin đang hoạt động"
    if "giohang" in s or "cart" in s or "dathang" in s:
        return "Đã đăng nhập, có sản phẩm trong giỏ hàng"
    if "sanpham" in s or "product" in s:
        return "Website đang hoạt động, có dữ liệu sản phẩm"
    return "Website đang hoạt động, trình duyệt đã mở"

def _make_postcondition(status, testcase):
    if status == "PASS":
        return "Hệ thống hoạt động đúng, không có lỗi server"
    return "Ghi nhận lỗi, cần kiểm tra lại"


# ──────────────────────────────────────────────────────────────
# Style helpers
# ──────────────────────────────────────────────────────────────
def _thin_border():
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _header_font():
    return Font(name="Arial", bold=True, color=COLOR_HEADER_FONT, size=10)

def _meta_font(bold=False):
    return Font(name="Arial", bold=bold, size=10)

def _cell_font():
    return Font(name="Arial", size=9)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _apply_header_row(ws, row_idx, values, col_widths=None):
    fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    for col_idx, val in enumerate(values, start=2):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = _header_font()
        cell.fill      = fill
        cell.alignment = _center()
        cell.border    = _thin_border()
    if col_widths:
        for col_idx, width in enumerate(col_widths, start=2):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_meta_block(ws, project, module, created_by, run_date, status_label):
    meta_fill = PatternFill("solid", fgColor=COLOR_META_BG)
    meta = [
        ("Project Name:",        project),
        ("Module Name:",         module),
        ("Created by:",          created_by),
        ("Date of run:",         run_date),
        ("Report type:",         status_label),
    ]
    for r, (label, value) in enumerate(meta, start=1):
        c_label = ws.cell(row=r, column=2, value=label)
        c_label.font = _meta_font(bold=True)
        c_label.fill = meta_fill
        c_label.alignment = _left()
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        c_val = ws.cell(row=r, column=3, value=value)
        c_val.font = _meta_font()
        c_val.fill = meta_fill
        c_val.alignment = _left()
    ws.row_dimensions[1].height = 18
    return len(meta) + 1


# ──────────────────────────────────────────────────────────────
# Columns — giống file mẫu
# ──────────────────────────────────────────────────────────────
COLUMNS = [
    "TEST CASE ID",
    "TEST SCENARIO",
    "TEST CASE",
    "PRE-CONDITION",
    "TEST STEPS",
    "TEST DATA",
    "EXPECTED RESULT",
    "POST CONDITION",
    "ACTUAL RESULT",
    "STATUS\n(PASS/FAIL)",
]
COL_WIDTHS = [18, 20, 32, 28, 40, 26, 32, 26, 32, 12]


def _parse_nodeid(nodeid):
    parts = nodeid.split("::")
    scenario  = parts[1] if len(parts) > 1 else ""
    func_full = parts[2] if len(parts) > 2 else (parts[1] if len(parts) > 1 else nodeid)
    m = re.match(r"^(.*?)\[(.+)\]$", func_full)
    if m:
        testcase = m.group(1)
        params   = m.group(2)
    else:
        testcase = func_full
        params   = ""
    return scenario, testcase, params


def _build_workbook(records, status_label, project, module, created_by, run_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.column_dimensions["A"].width = 2

    next_row = _apply_meta_block(ws, project, module, created_by, run_date, status_label)
    next_row += 1

    header_row = next_row
    _apply_header_row(ws, header_row, COLUMNS, COL_WIDTHS)
    ws.row_dimensions[header_row].height = 32
    next_row += 1

    pass_fill = PatternFill("solid", fgColor=COLOR_PASS_BG)
    fail_fill = PatternFill("solid", fgColor=COLOR_FAIL_BG)
    alt_fill  = PatternFill("solid", fgColor=COLOR_ROW_ALT)

    for idx, rec in enumerate(records, start=1):
        r = next_row + idx - 1
        ws.row_dimensions[r].height = 52   # cao hơn vì có test steps

        is_pass  = rec["status"] == "PASS"
        row_fill = (alt_fill if is_pass and idx % 2 == 0 else
                    pass_fill if is_pass else fail_fill)

        prefix = _get_prefix(rec["nodeid"])
        scenario, testcase, params = _parse_nodeid(rec["nodeid"])

        pre   = _make_precondition(scenario)
        steps = _make_steps(testcase, params)
        post  = _make_postcondition(rec["status"], testcase)

        expected = rec.get("expected") or "Không có lỗi server / hoạt động đúng nghiệp vụ"
        actual   = rec.get("actual")   or ("PASS - Không có lỗi" if is_pass
                                           else f"FAIL - {rec.get('error','')[:80]}")

        values = [
            _tc_id(prefix, idx),   # TEST CASE ID
            scenario,              # TEST SCENARIO
            testcase,              # TEST CASE
            pre,                   # PRE-CONDITION
            steps,                 # TEST STEPS
            params,                # TEST DATA
            expected,              # EXPECTED RESULT
            post,                  # POST CONDITION
            actual,                # ACTUAL RESULT
            rec["status"],         # STATUS
        ]

        for col_idx, val in enumerate(values, start=2):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font      = _cell_font()
            cell.fill      = row_fill
            cell.border    = _thin_border()
            # STATUS và ID căn giữa, còn lại căn trái
            cell.alignment = _center() if col_idx in (2, 11) else _left()

            # STATUS cell: in đậm màu chữ
            if col_idx == 11:
                cell.font = Font(name="Arial", bold=True,
                                 color="006400" if is_pass else "8B0000", size=9)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 2
    ws_sum.column_dimensions["B"].width = 30
    ws_sum.column_dimensions["C"].width = 16
    fill_sum = PatternFill("solid", fgColor=COLOR_META_BG)
    total    = len(records)
    passed   = sum(1 for r in records if r["status"] == "PASS")
    failed   = total - passed
    for r_idx, (k, v) in enumerate([
        ("Report type",   status_label),
        ("Total",         total),
        ("PASS",          passed),
        ("FAIL",          failed),
        ("Generated",     run_date),
    ], start=2):
        c1 = ws_sum.cell(row=r_idx, column=2, value=k)
        c1.font = _meta_font(bold=True); c1.fill = fill_sum; c1.alignment = _left()
        c2 = ws_sum.cell(row=r_idx, column=3, value=v)
        c2.font = _meta_font(); c2.fill = fill_sum; c2.alignment = _left()

    return wb


# ──────────────────────────────────────────────────────────────
# Pytest hooks
# ──────────────────────────────────────────────────────────────
_results_store = []


def pytest_runtest_logreport(report):
    if report.when != "call":
        return
    duration  = getattr(report, "duration", 0.0)
    error_msg = ""
    if report.failed:
        if hasattr(report, "longrepr"):
            lines = str(report.longrepr).splitlines()
            meaningful = [l for l in lines if l.strip()]
            error_msg = meaningful[-1] if meaningful else str(report.longrepr)[:200]

    _results_store.append({
        "nodeid":   report.nodeid,
        "status":   "PASS" if report.passed else "FAIL",
        "duration": duration,
        "error":    error_msg,
        "expected": "Không có lỗi server / hoạt động đúng nghiệp vụ",
        "actual":   "PASS - Không có lỗi" if report.passed else f"FAIL - {error_msg[:80]}",
    })


def pytest_sessionfinish(session, exitstatus):
    if not _results_store:
        return

    now        = datetime.datetime.now()
    run_date   = now.strftime("%d/%m/%Y %H:%M:%S")
    project    = "Jollibee Web - ASP.NET MVC"
    module     = "Selenium Automation Test"
    created_by = "Phạm Ánh Ngọc - 2001230566"
    out_dir    = os.path.join(os.getcwd(), "test_reports")
    os.makedirs(out_dir, exist_ok=True)

    passed = [r for r in _results_store if r["status"] == "PASS"]
    failed = [r for r in _results_store if r["status"] == "FAIL"]

    path_pass = os.path.join(out_dir, "test_results_PASS.xlsx")
    path_fail = os.path.join(out_dir, "test_results_FAIL.xlsx")

    if passed:
        wb = _build_workbook(passed, "✅ PASS ONLY", project, module, created_by, run_date)
        wb.save(path_pass)
        print(f"\n✅  PASS report → {path_pass}  ({len(passed)} tests)")

    if failed:
        wb = _build_workbook(failed, "❌ FAIL ONLY", project, module, created_by, run_date)
        wb.save(path_fail)
        print(f"❌  FAIL report → {path_fail}  ({len(failed)} tests)")

    if not passed and not failed:
        print("\n⚠️  Không có kết quả để xuất (có thể tất cả bị skip).")